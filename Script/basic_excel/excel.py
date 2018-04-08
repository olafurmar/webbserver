# -*- coding: utf-8 -*-
"""
Module basic_excel
==================

Implements a few basic openpyxl functions that work for both .xls and .xlsx
files, using openpyxl and xlrd as needed. For more advanced use, the
appropriate package should be used instead.

Created on Wed Dec 13 14:24:31 2017

@author: fkln
"""

import datetime as dt
import openpyxl
import os
import xlrd

from openpyxl.utils.exceptions import InvalidFileException
from xlrd.biffh import XLRDError
from xlutils.copy import copy
from xlutils.compat import BytesIO
from zipfile import BadZipfile



def load_workbook(filename):
    '''Returns a workbook object.
    Accepts a path to an excel file as input.
    '''
    
    try:
        # First try to open, using Openpyxl
        wb = openpyxl.load_workbook(filename = filename)
        return Workbook(wb, legacy = False)
    except InvalidFileException:
        # Wrong format - try using xlrd instead
        wb = xlrd.open_workbook(filename = filename)
        return Workbook(wb, legacy = True)
    except (IOError, BadZipfile):
        # File like object of wrong format used?
        try:
            filename.seek(0)
        except AttributeError:
            raise TypeError('Unknown input format.')
        wb = xlrd.open_workbook(file_contents = filename.read())
        return Workbook(wb, legacy = True)
        

class DocProps(object):
    def __init__(self, modified):
        if isinstance(modified, dt.datetime):
            self.modified = modified
        else:
            raise TypeError('modified date not of type datetime.')

class Workbook(object):
    '''Implements a basic Excel workbook object.
    '''
    def __init__(self, wb_object, legacy):
        self._legacy = legacy
        self.wb = wb_object
        if legacy:
            # Get active sheet
            for sht in wb_object.sheets():
                if sht.sheet_visible:
                    # This is the active sheet (there could be multiple but
                    # we will get the first one)
                    self.active = Worksheet(sht, self, legacy = True)
                    break
            # Date is currently not supported for legacy files - always set to today
            self.properties = DocProps(modified = dt.datetime.today())
        else:
            self.active = Worksheet(wb_object.active, self, legacy = False)
            self.properties = DocProps(modified=wb_object.properties.modified)
            
    def get_sheet_by_name(self, sheet_name):
        '''Returns the worksheet with the specified name.
        '''
        if self._legacy:
            try:
                return Worksheet(self.wb.sheet_by_name(sheet_name), self, legacy = True)
            except XLRDError:
                # The sheet does not exist
                raise KeyError(u'Worksheet %s does not exist.' % sheet_name)
        else:
            return Worksheet(self.wb.get_sheet_by_name(sheet_name), self, legacy = False)
        
    def get_sheet_names(self):
        if self._legacy:
            return self.wb.sheet_names()
        else:
            return self.wb.get_sheet_names()
    
    def save(self, filename):
        '''Saves the workbook with the provided filename.
        '''
        if self._legacy:
            # Get an xlwt version of the workbook and save it
            wt_wb = copy(self.wb)
            wt_wb.save(filename)
            del wt_wb
        else:
            self.wb.save(filename)

class Worksheet(object):
    '''Implements a basic Excel worksheet object.
    '''
    def __init__(self, ws_object, basic_wb, legacy):
        self.ws = ws_object
        self.parent = basic_wb
        self._legacy = legacy
        if legacy:
            self.max_row = ws_object.nrows
        else:
            self.max_row = ws_object.max_row
            
    def cell(self, row, column, value = None):
        '''Without value: Returns the cell on row, column as defined in Openpyxl.
        With value: Sets the value of the specified cell.
        '''
        if self._legacy:
            if value == None:
                return Cell(value = self.ws.cell_value(rowx = row-1, colx = column-1), legacy = True)
            else:
                # Editing a legacy sheet is quite complicated
                # First note the name of the current sheet
                sheet_name = self.ws.name
                # Get an xlwt version of the workbook
                wt_wb = copy(self.parent.wb)
                # Now get the same sheet again
                wt_sheet = wt_wb.get_sheet(sheet_name)
                # Write to the cell
                wt_sheet.write(row-1,column-1,value)
                # Now clean up and read the workbook back
                s = BytesIO()
                wt_wb.save(s)
                s.seek(0)
                rd_wb = xlrd.open_workbook(file_contents = s.read())
                self.parent.__init__(rd_wb, legacy = True)
                self.__init__(self.parent.get_sheet_by_name(sheet_name).ws, self.parent, legacy=True)
                del wt_wb, wt_sheet, s, sheet_name
        else:
            if value == None:
                return Cell(cell_object = self.ws.cell(row = row, column = column), legacy = False)
            else:
                self.ws.cell(row = row, column = column, value = value)
                self.parent.__init__(self.parent.wb,legacy = False)
                self.__init__(self.ws, self.parent, legacy=False)

class Cell(object):
    '''Implements a basic Excel cell object.
    '''
    def __init__(self, cell_object = None, value = None, legacy = False):
        if cell_object != None and value != None:
            raise ValueError('Only one of cell_object and value can be specified at any time.')
        elif cell_object != None:
            self.cell = cell_object
            self.value = cell_object.value
        elif value != None:
            self.value = value
        else:
            raise ValueError('One of cell_object and value shall be specified when calling Cell.')
        
        self._legacy = legacy
        
if __name__ == '__main__':
    pass