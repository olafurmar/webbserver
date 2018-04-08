# -*- coding: utf-8 -*-
"""
Tests of the excel module.

Created on Wed Dec 13 14:28:39 2017

@author: fkln
"""

import excel
import openpyxl
import pytest
import shutil
import xlrd

from os.path import abspath, basename, isfile, join
from arena.psr2arena.utils import add_before_file_ext

@pytest.fixture(scope='module')
def legacy_file(tmpdir_factory):
    '''A workbook saved in .xls format.
    '''
    src_file = u'test_resources\\legacy.xls'
    dest_file = join(str(tmpdir_factory.getbasetemp()),
                             basename(src_file))
    shutil.copy(abspath(src_file), dest_file)
    return dest_file

@pytest.fixture(scope='module')
def current_file(tmpdir_factory):
    '''The same workbook as legacy_file, but saved in .xlsx format.
    '''
    src_file = u'test_resources\\current.xlsx'
    dest_file = join(str(tmpdir_factory.getbasetemp()),
                             basename(src_file))
    shutil.copy(abspath(src_file), dest_file)
    return dest_file


def test_load_workbook(legacy_file, current_file):
    '''Tests load_workbook function.
    '''
    for in_file in [legacy_file, current_file]:
        wb = excel.load_workbook(in_file)
        assert isinstance(wb, excel.Workbook)
        
def test_get_active_worksheet(legacy_file, current_file):
    '''Tests that the active worksheet is available as an attribute
    on the workbook.
    '''
    for in_file in [legacy_file, current_file]:
        wb = excel.load_workbook(in_file)
        aws = wb.active
        assert isinstance(aws, excel.Worksheet)

def test_get_sheet_by_name_and_cell_value_by_rc(legacy_file, current_file):
    '''Tests that it is possible to get a sheet by name and to get the
    value of a cell by row and column. Uses 1,1 as the first cell (as in 
    Openpyxl).
    '''
    for in_file in [legacy_file, current_file]:
        wb = excel.load_workbook(in_file)
        ws = wb.get_sheet_by_name('test_sheet')
        assert ws.cell(row = 1, column = 1).value == u'ÄBC'

def test_get_wb_from_ws(legacy_file, current_file):
    '''Tests that the workbook is available as the .parent property
    on the worksheet.
    '''
    for in_file in [legacy_file, current_file]:
        wb = excel.load_workbook(in_file)
        aws = wb.active
        assert wb == aws.parent

def test_save_wb(legacy_file, current_file, tmpdir_factory):
    '''Tests that the workbook can be saved as in openpyxl.
    '''
    for in_file in [legacy_file, current_file]:
        wb = excel.load_workbook(in_file)
        dest_file = add_before_file_ext(join(str(tmpdir_factory.getbasetemp()),
                         basename(in_file)), 'test01')
        wb.save(dest_file)
        assert isfile(dest_file)
        
def test_pass_filelike_to_load_workbook(legacy_file, current_file):
    '''Tests that it is possible to pass the workbook as a string to
    load_workbook.
    '''
    for in_file in [legacy_file, current_file]:
        with open(in_file, 'rb') as f:
            wb = excel.load_workbook(f)
        assert wb.active.cell(row = 1, column = 1).value == u'Adam'

def test_get_max_row(legacy_file, current_file):
    '''Tests that it is possible to get the last row of the sheet.
    '''
    sheet_name = u'Sheet1'
    # Get the right answer
    max_row = openpyxl.load_workbook(current_file).get_sheet_by_name(sheet_name).max_row
    for in_file in [legacy_file, current_file]:
        wb = excel.load_workbook(in_file)
        ws = wb.get_sheet_by_name(sheet_name)
        ws_max_row = ws.max_row
        assert ws_max_row == max_row
        # Now add a cell beyond the last row
        new_last = ws_max_row + 10
        ws.cell(row = new_last, column = 1, value = u'ÅÄÖ')
        assert ws.max_row == new_last

if __name__ == '__main__':
    pytest.main([r'--basetemp=c:\test_excel', r'test_excel.py'])
    #pytest.main([r'--basetemp=c:\test_excel', '-k', 'test_get_max_row', r'test_excel.py'])
